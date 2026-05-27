"""Bulk XLSX import for Bayaan master data — products, suppliers, ingredients, recipes.

The client onboards 10–50 kiosks of master data; one-record-at-a-time JSON API is
not realistic. This controller exposes:

  GET  /bayaan/api/import/template/<entity>
       → Pre-formatted XLSX with header row + one example row + inline comments.

  POST /bayaan/api/import/<entity>
       Body: { "filename": str, "content_b64": str, "commit": bool }
       Default is dry-run: parses, validates, returns row counts + errors but
       does not write to the database. With "commit": true the route writes in
       a single transaction; if any row fails validation, nothing commits.

Manager-only access enforced server-side. Idempotency: rows are matched by an
external_ref column (or the natural key for the entity) so re-importing the same
file does not duplicate records.
"""

import base64
import io
import logging
from collections import OrderedDict

from odoo import http
from odoo.exceptions import UserError
from odoo.http import request, Response

_logger = logging.getLogger(__name__)


# Each entity defines: name, key column, columns spec, example row, importer function.
# The columns spec is ordered: header → column name (str) + optional comment.

PRODUCT_COLUMNS = OrderedDict([
    ("external_ref", "Stable code you control; re-importing the same code updates the existing record."),
    ("name", "Display name (English)."),
    ("name_ar", "Display name (Arabic). Optional but recommended for receipts."),
    ("category", "POS category, e.g. Hot Coffee, Iced Coffee, Juice, Cake, Bakery."),
    ("price", "Sell price in IQD."),
    ("consumption_mode", "One of: recipe, finished, hybrid, none. Decides whether Bayaan deducts ingredients."),
    ("default_size", "S, M, L, slice, pc — the default size shown in POS."),
    ("active", "TRUE/FALSE — leave blank for TRUE."),
])
PRODUCT_EXAMPLE = ["PRD-COFFEE-LATTE", "Latte", "لاتيه", "Hot Coffee", 5000, "recipe", "M", "TRUE"]

SUPPLIER_COLUMNS = OrderedDict([
    ("external_ref", "Stable supplier code, e.g. SUP-BAGHDAD-DAIRY."),
    ("name", "Display name."),
    ("category", "Free text category like Produce, Dairy, Packaging."),
    ("phone", "Contact phone (any format)."),
    ("email", "Contact email."),
    ("address", "Single-line postal address."),
    ("vat_id", "Tax/registration number if applicable."),
    ("active", "TRUE/FALSE — leave blank for TRUE."),
])
SUPPLIER_EXAMPLE = ["SUP-BAGHDAD-DAIRY", "Baghdad Dairy", "Dairy", "+9647700000000", "ops@baghdad-dairy.iq", "Karrada Industrial Zone, Baghdad", "1234567", "TRUE"]

INGREDIENT_COLUMNS = OrderedDict([
    ("external_ref", "Stable ingredient code, e.g. ING-WHOLE-MILK."),
    ("name", "Display name."),
    ("uom", "Unit code, e.g. kg, L, pc, g, ml. Must already exist in Odoo as uom.uom."),
    ("supplier_ref", "Default supplier external_ref. Optional."),
    ("standard_cost", "Latest unit cost in IQD. Used for valuation/landed costs."),
    ("reorder_qty", "Reorder when on-hand drops below this quantity."),
    ("max_qty", "Suggested order-up-to quantity."),
])
INGREDIENT_EXAMPLE = ["ING-WHOLE-MILK", "Whole milk", "L", "SUP-BAGHDAD-DAIRY", 1250, 80, 200]

RECIPE_COLUMNS = OrderedDict([
    ("product_ref", "Product external_ref this recipe applies to."),
    ("ingredient_ref", "Ingredient external_ref."),
    ("qty_per_unit", "Quantity of ingredient consumed per one finished unit sold."),
    ("uom", "UoM used in the qty column. Must match the ingredient UoM."),
    ("effective_from", "ISO date YYYY-MM-DD. Recipe activates from this date forward."),
])
RECIPE_EXAMPLE = ["PRD-COFFEE-LATTE", "ING-WHOLE-MILK", 0.22, "L", "2026-05-21"]


ENTITY_SPECS = {
    "products": {
        "label": "Products / Menu",
        "columns": PRODUCT_COLUMNS,
        "example": PRODUCT_EXAMPLE,
    },
    "suppliers": {
        "label": "Suppliers",
        "columns": SUPPLIER_COLUMNS,
        "example": SUPPLIER_EXAMPLE,
    },
    "ingredients": {
        "label": "Ingredients / Stock items",
        "columns": INGREDIENT_COLUMNS,
        "example": INGREDIENT_EXAMPLE,
    },
    "recipes": {
        "label": "Recipes (BOM per product)",
        "columns": RECIPE_COLUMNS,
        "example": RECIPE_EXAMPLE,
    },
}


def _is_bayaan_manager_request():
    user = request.env.user
    return user.has_group("base.group_system") or user.has_group(
        "bayaan_fnb_kiosk.group_bayaan_manager"
    )


def _build_template(entity_key):
    spec = ENTITY_SPECS[entity_key]
    try:
        import openpyxl
        from openpyxl.comments import Comment
    except ImportError as exc:  # pragma: no cover — openpyxl ships with Odoo deps
        raise UserError("openpyxl is required for XLSX templates: %s" % exc)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = entity_key
    headers = list(spec["columns"].keys())
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        comment_text = spec["columns"][header]
        if comment_text:
            cell.comment = Comment(comment_text, "Bayaan import")
        ws.column_dimensions[cell.column_letter].width = max(18, len(header) + 6)
    # Example row.
    for col_idx, value in enumerate(spec["example"], start=1):
        ws.cell(row=2, column=col_idx, value=value)
    # Help row.
    ws.cell(row=3, column=1, value="↑ Row 2 is an example — delete it before import or overwrite with real data.")
    ws.cell(row=3, column=1).font = openpyxl.styles.Font(italic=True, color="808080")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_xlsx_rows(content_b64, expected_columns):
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise UserError("openpyxl is required for XLSX import: %s" % exc)
    try:
        raw = base64.b64decode(content_b64 or "", validate=False)
    except Exception as exc:
        raise UserError("content_b64 is not valid base64: %s" % exc)
    if not raw:
        raise UserError("Empty XLSX payload.")
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as exc:
        raise UserError("Could not parse XLSX: %s" % exc)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise UserError("XLSX is empty — no header row.")
    headers = [(str(h).strip() if h is not None else "") for h in header_row]
    missing = [col for col in expected_columns if col not in headers]
    if missing:
        raise UserError(
            "XLSX is missing required columns: %s. Download the template to see the expected schema."
            % ", ".join(missing)
        )
    data_rows = []
    for r_idx, raw_row in enumerate(rows_iter, start=2):
        record = {headers[i]: raw_row[i] for i in range(len(headers)) if i < len(raw_row)}
        # Skip blank rows or example-only rows that start with ↑.
        if not any(v not in (None, "", " ") for v in record.values()):
            continue
        if any(isinstance(v, str) and v.strip().startswith("↑") for v in record.values()):
            continue
        record["_row_number"] = r_idx
        data_rows.append(record)
    return data_rows


def _truthy(value, default=True):
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in ("true", "1", "yes", "y", "نعم")


def _import_suppliers(env, rows, commit):
    Partner = env["res.partner"].sudo()
    created, updated, errors = [], [], []
    for row in rows:
        try:
            external_ref = (row.get("external_ref") or "").strip()
            name = (row.get("name") or "").strip()
            if not external_ref:
                raise ValueError("external_ref is required")
            if not name:
                raise ValueError("name is required")
            vals = {
                "name": name,
                "supplier_rank": 1,
                "active": _truthy(row.get("active"), True),
                "phone": (row.get("phone") or "") or False,
                "email": (row.get("email") or "") or False,
                "street": (row.get("address") or "") or False,
                "vat": (row.get("vat_id") or "") or False,
                "ref": external_ref,
                "comment": ("Category: %s" % row.get("category")) if row.get("category") else False,
            }
            existing = Partner.search([("ref", "=", external_ref)], limit=1)
            if existing:
                if commit:
                    existing.write(vals)
                updated.append(external_ref)
            else:
                if commit:
                    Partner.create(vals)
                created.append(external_ref)
        except Exception as exc:
            errors.append({"row": row.get("_row_number"), "ref": row.get("external_ref"), "error": str(exc)})
    return created, updated, errors


def _import_ingredients(env, rows, commit):
    Product = env["product.product"].sudo()
    Uom = env["uom.uom"].sudo()
    Partner = env["res.partner"].sudo()
    created, updated, errors = [], [], []
    for row in rows:
        try:
            external_ref = (row.get("external_ref") or "").strip()
            name = (row.get("name") or "").strip()
            uom_code = (row.get("uom") or "").strip().lower()
            if not external_ref:
                raise ValueError("external_ref is required")
            if not name:
                raise ValueError("name is required")
            uom = False
            if uom_code:
                uom = Uom.search([("name", "=ilike", uom_code)], limit=1)
                if not uom:
                    raise ValueError("uom %r not found in Odoo" % uom_code)
            supplier = False
            supplier_ref = (row.get("supplier_ref") or "").strip()
            if supplier_ref:
                supplier = Partner.search([("ref", "=", supplier_ref), ("supplier_rank", ">", 0)], limit=1)
                if not supplier:
                    raise ValueError("supplier_ref %r not found — import suppliers first" % supplier_ref)
            try:
                standard_cost = float(row.get("standard_cost") or 0)
            except Exception:
                raise ValueError("standard_cost must be numeric")
            vals = {
                "name": name,
                "type": "consu",
                "is_storable": True,
                "default_code": external_ref,
                "standard_price": standard_cost,
            }
            if uom:
                vals["uom_id"] = uom.id
                vals["uom_po_id"] = uom.id
            existing = Product.search([("default_code", "=", external_ref)], limit=1)
            if existing:
                if commit:
                    existing.write(vals)
                updated.append(external_ref)
            else:
                if commit:
                    Product.create(vals)
                created.append(external_ref)
        except Exception as exc:
            errors.append({"row": row.get("_row_number"), "ref": row.get("external_ref"), "error": str(exc)})
    return created, updated, errors


def _import_products(env, rows, commit):
    Product = env["product.product"].sudo()
    PosCategory = env["pos.category"].sudo()
    created, updated, errors = [], [], []
    valid_modes = {"recipe", "finished", "hybrid", "none"}
    for row in rows:
        try:
            external_ref = (row.get("external_ref") or "").strip()
            name = (row.get("name") or "").strip()
            if not external_ref:
                raise ValueError("external_ref is required")
            if not name:
                raise ValueError("name is required")
            mode = (row.get("consumption_mode") or "recipe").strip().lower()
            if mode not in valid_modes:
                raise ValueError("consumption_mode must be one of %s" % sorted(valid_modes))
            try:
                price = float(row.get("price") or 0)
            except Exception:
                raise ValueError("price must be numeric")
            if price < 0:
                raise ValueError("price must be >= 0")
            pos_category = False
            category_name = (row.get("category") or "").strip()
            if category_name:
                pos_category = PosCategory.search([("name", "=", category_name)], limit=1)
                if not pos_category and commit:
                    pos_category = PosCategory.create({"name": category_name})
            vals = {
                "name": name,
                "type": "consu",
                "default_code": external_ref,
                "lst_price": price,
                "available_in_pos": True,
                "active": _truthy(row.get("active"), True),
                "bayaan_consumption_mode": mode,
            }
            if pos_category:
                vals["pos_categ_ids"] = [(6, 0, [pos_category.id])]
            existing = Product.search([("default_code", "=", external_ref)], limit=1)
            if existing:
                if commit:
                    existing.write(vals)
                updated.append(external_ref)
            else:
                if commit:
                    Product.create(vals)
                created.append(external_ref)
        except Exception as exc:
            errors.append({"row": row.get("_row_number"), "ref": row.get("external_ref"), "error": str(exc)})
    return created, updated, errors


def _import_recipes(env, rows, commit):
    """Group rows by product_ref + effective_from and create bayaan.recipe with lines."""
    Product = env["product.product"].sudo()
    Recipe = env["bayaan.recipe"].sudo()
    Uom = env["uom.uom"].sudo()
    created, updated, errors = [], [], []
    grouped = {}
    for row in rows:
        product_ref = (row.get("product_ref") or "").strip()
        effective_from = (row.get("effective_from") or "").strip() or None
        key = (product_ref, effective_from)
        grouped.setdefault(key, []).append(row)
    for (product_ref, effective_from), group_rows in grouped.items():
        try:
            if not product_ref:
                raise ValueError("product_ref is required")
            product = Product.search([("default_code", "=", product_ref)], limit=1)
            if not product:
                raise ValueError("product_ref %r not found — import the product first" % product_ref)
            line_payloads = []
            for r in group_rows:
                ing_ref = (r.get("ingredient_ref") or "").strip()
                if not ing_ref:
                    raise ValueError("ingredient_ref is required on row %s" % r.get("_row_number"))
                ing = Product.search([("default_code", "=", ing_ref)], limit=1)
                if not ing:
                    raise ValueError("ingredient_ref %r not found — import ingredients first" % ing_ref)
                try:
                    qty = float(r.get("qty_per_unit") or 0)
                except Exception:
                    raise ValueError("qty_per_unit must be numeric on row %s" % r.get("_row_number"))
                if qty <= 0:
                    raise ValueError("qty_per_unit must be > 0 on row %s" % r.get("_row_number"))
                uom_code = (r.get("uom") or "").strip()
                uom = False
                if uom_code:
                    uom = Uom.search([("name", "=ilike", uom_code)], limit=1)
                    if not uom:
                        raise ValueError("uom %r not found on row %s" % (uom_code, r.get("_row_number")))
                else:
                    uom = ing.uom_id
                if not uom:
                    raise ValueError("uom is required for ingredient on row %s" % r.get("_row_number"))
                line_payloads.append({"ingredient_id": ing.id, "qty": qty, "uom_id": uom.id})
            domain = [("product_id", "=", product.id)]
            if effective_from:
                domain.append(("effective_from", "=", effective_from))
            existing = Recipe.search(domain, limit=1, order="effective_from desc")
            if existing:
                if commit:
                    existing.line_ids.unlink()
                    existing.write({"line_ids": [(0, 0, p) for p in line_payloads]})
                updated.append("%s @ %s" % (product_ref, effective_from or "latest"))
            else:
                if commit:
                    vals = {
                        "product_id": product.id,
                        "line_ids": [(0, 0, p) for p in line_payloads],
                    }
                    if effective_from:
                        vals["effective_from"] = effective_from
                    Recipe.create(vals)
                created.append("%s @ %s" % (product_ref, effective_from or "latest"))
        except Exception as exc:
            errors.append({
                "product_ref": product_ref,
                "effective_from": effective_from,
                "error": str(exc),
            })
    return created, updated, errors


IMPORTERS = {
    "products": _import_products,
    "suppliers": _import_suppliers,
    "ingredients": _import_ingredients,
    "recipes": _import_recipes,
}


class BayaanDataImportController(http.Controller):
    @http.route("/bayaan/api/import/template/<string:entity>", type="http", auth="user", methods=["GET"], csrf=False)
    def download_template(self, entity, **_kw):
        if not _is_bayaan_manager_request():
            raise UserError("Only Bayaan managers can download import templates.")
        if entity not in ENTITY_SPECS:
            raise UserError("Unknown import entity %r" % entity)
        payload = _build_template(entity)
        filename = "bayaan-%s-template.xlsx" % entity
        return Response(
            payload,
            headers=[
                ("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ("Content-Disposition", "attachment; filename=%s" % filename),
                ("Content-Length", str(len(payload))),
            ],
        )

    @http.route("/bayaan/api/import/<string:entity>", type="json", auth="user", methods=["POST"], csrf=False)
    def import_entity(self, entity, **payload):
        if not _is_bayaan_manager_request():
            raise UserError("Only Bayaan managers can run master-data imports.")
        if entity not in ENTITY_SPECS:
            raise UserError("Unknown import entity %r" % entity)
        # Payload may be passed flat or under a 'payload' key (Bayaan JSON-RPC convention).
        body = payload.get("payload", payload) or {}
        content_b64 = body.get("content_b64") or ""
        commit = bool(body.get("commit"))
        spec = ENTITY_SPECS[entity]
        rows = _parse_xlsx_rows(content_b64, list(spec["columns"].keys()))
        importer = IMPORTERS[entity]
        if not commit:
            # Dry-run inside a savepoint that always rolls back.
            env = request.env
            try:
                with env.cr.savepoint():
                    created, updated, errors = importer(env, rows, commit=True)
                    raise _DryRunRollback()
            except _DryRunRollback:
                pass
        else:
            env = request.env
            created, updated, errors = importer(env, rows, commit=True)
            if errors:
                # Any error in commit mode rolls back the whole transaction so the
                # client sees a consistent before/after state.
                raise UserError(
                    "Import failed for %d row(s); nothing committed. First error: %s"
                    % (len(errors), errors[0])
                )
        return {
            "ok": not errors,
            "entity": entity,
            "rows_seen": len(rows),
            "created": len(created),
            "updated": len(updated),
            "errors": errors,
            "committed": commit and not errors,
            "created_refs": created,
            "updated_refs": updated,
        }


class _DryRunRollback(Exception):
    """Raised inside a savepoint to roll back a dry-run import."""

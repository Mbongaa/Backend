import base64
import io

from odoo.tests.common import HttpCase, tagged

from .common import BayaanTestBase


def _build_xlsx(headers, rows):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode("ascii")


@tagged("post_install", "-at_install")
class TestDataImport(BayaanTestBase, HttpCase):
    """Master-data XLSX import: dry-run vs commit, role enforcement, validation."""

    def _jsonrpc(self, route, payload):
        return self.opener.post(
            self.base_url() + route,
            json={"jsonrpc": "2.0", "method": "call", "params": {"payload": payload}, "id": 1},
        ).json()

    def setUp(self):
        super().setUp()
        self.authenticate("admin", "admin")
        manager_group = self.env.ref("bayaan_fnb_kiosk.group_bayaan_manager")
        self.manager = self.env["res.users"].sudo().create({
            "name": "import_mgr",
            "login": "bayaan_import_manager",
            "password": "test",
            "company_id": self.company.id,
            "company_ids": [(6, 0, [self.company.id])],
            "group_ids": [(6, 0, [manager_group.id])],
        })
        cashier_group = self.env.ref("bayaan_fnb_kiosk.group_bayaan_cashier")
        self.cashier = self.env["res.users"].sudo().create({
            "name": "import_cashier",
            "login": "bayaan_import_cashier",
            "password": "test",
            "company_id": self.company.id,
            "company_ids": [(6, 0, [self.company.id])],
            "group_ids": [(6, 0, [cashier_group.id])],
        })

    def test_template_download_for_manager(self):
        self.authenticate("bayaan_import_manager", "test")
        response = self.opener.get(self.base_url() + "/bayaan/api/import/template/suppliers")
        self.assertEqual(response.status_code, 200)
        self.assertIn("application/vnd.openxmlformats", response.headers.get("Content-Type", ""))
        self.assertGreater(len(response.content), 1000)

    def test_template_download_blocked_for_cashier(self):
        self.authenticate("bayaan_import_cashier", "test")
        response = self.opener.get(self.base_url() + "/bayaan/api/import/template/suppliers")
        # The route raises UserError; HTTP returns 200 with an error page or 500 depending on Odoo
        # — what we really care about is the cashier cannot pull a templated file.
        self.assertNotIn("application/vnd.openxmlformats", response.headers.get("Content-Type", ""))

    def test_dry_run_reports_rows_without_writing(self):
        self.authenticate("bayaan_import_manager", "test")
        xlsx = _build_xlsx(
            ["external_ref", "name", "category", "phone", "email", "address", "vat_id", "active"],
            [
                ["SUP-DRYRUN-1", "Dry Run Supplier", "Produce", "111", "a@x.test", "addr", "VAT1", "TRUE"],
                ["SUP-DRYRUN-2", "Another Dry Run", "Dairy", "222", "b@x.test", "addr2", "VAT2", "TRUE"],
            ],
        )
        response = self._jsonrpc("/bayaan/api/import/suppliers", {"content_b64": xlsx, "commit": False})
        result = response.get("result")
        self.assertIsNotNone(result, msg=str(response))
        self.assertEqual(result["rows_seen"], 2)
        self.assertEqual(result["created"], 2)
        self.assertFalse(result["committed"])
        # Confirm DB was NOT mutated.
        self.assertFalse(self.env["res.partner"].sudo().search([("ref", "=", "SUP-DRYRUN-1")]))

    def test_commit_creates_then_updates_idempotently(self):
        self.authenticate("bayaan_import_manager", "test")
        xlsx = _build_xlsx(
            ["external_ref", "name", "category", "phone", "email", "address", "vat_id", "active"],
            [
                ["SUP-CREATE-1", "Initial Name", "Produce", "111", "a@x.test", "addr", "VAT1", "TRUE"],
            ],
        )
        first = self._jsonrpc("/bayaan/api/import/suppliers", {"content_b64": xlsx, "commit": True})
        self.assertEqual(first["result"]["created"], 1)
        self.assertEqual(first["result"]["updated"], 0)
        self.assertTrue(first["result"]["committed"])
        partner = self.env["res.partner"].sudo().search([("ref", "=", "SUP-CREATE-1")], limit=1)
        self.assertTrue(partner)
        self.assertEqual(partner.name, "Initial Name")
        # Re-import same ref with updated name → update, not duplicate.
        xlsx2 = _build_xlsx(
            ["external_ref", "name", "category", "phone", "email", "address", "vat_id", "active"],
            [
                ["SUP-CREATE-1", "Renamed Co", "Dairy", "999", "z@x.test", "addr2", "VAT2", "TRUE"],
            ],
        )
        second = self._jsonrpc("/bayaan/api/import/suppliers", {"content_b64": xlsx2, "commit": True})
        self.assertEqual(second["result"]["created"], 0)
        self.assertEqual(second["result"]["updated"], 1)
        partners = self.env["res.partner"].sudo().search([("ref", "=", "SUP-CREATE-1")])
        self.assertEqual(len(partners), 1)
        self.assertEqual(partners.name, "Renamed Co")

    def test_missing_required_column_rejected(self):
        self.authenticate("bayaan_import_manager", "test")
        xlsx = _build_xlsx(
            ["external_ref", "name"],  # missing many required columns
            [["SUP-X", "X"]],
        )
        response = self._jsonrpc("/bayaan/api/import/suppliers", {"content_b64": xlsx, "commit": False})
        self.assertIn("error", response, msg=str(response))
        self.assertIn("missing required columns", str(response["error"]).lower())

    def test_blank_external_ref_row_recorded_as_error(self):
        self.authenticate("bayaan_import_manager", "test")
        xlsx = _build_xlsx(
            ["external_ref", "name", "category", "phone", "email", "address", "vat_id", "active"],
            [["", "No-Ref Co", "X", "", "", "", "", "TRUE"]],
        )
        response = self._jsonrpc("/bayaan/api/import/suppliers", {"content_b64": xlsx, "commit": False})
        result = response.get("result")
        self.assertIsNotNone(result, msg=str(response))
        self.assertEqual(len(result["errors"]), 1)
        self.assertIn("external_ref", result["errors"][0]["error"].lower())

    def test_commit_rolls_back_on_any_error(self):
        self.authenticate("bayaan_import_manager", "test")
        xlsx = _build_xlsx(
            ["external_ref", "name", "category", "phone", "email", "address", "vat_id", "active"],
            [
                ["SUP-ROLLBACK-OK", "OK Co", "X", "", "", "", "", "TRUE"],
                ["", "No-Ref Co", "X", "", "", "", "", "TRUE"],
            ],
        )
        response = self._jsonrpc("/bayaan/api/import/suppliers", {"content_b64": xlsx, "commit": True})
        # commit=True + any error → UserError surfaced through JSON-RPC.
        self.assertIn("error", response, msg=str(response))
        # The OK row must NOT have landed (savepoint or transaction rollback).
        self.assertFalse(self.env["res.partner"].sudo().search([("ref", "=", "SUP-ROLLBACK-OK")]))

    def test_cashier_cannot_import(self):
        self.authenticate("bayaan_import_cashier", "test")
        xlsx = _build_xlsx(
            ["external_ref", "name", "category", "phone", "email", "address", "vat_id", "active"],
            [["SUP-BLOCK", "Blocked", "Produce", "", "", "", "", "TRUE"]],
        )
        response = self._jsonrpc("/bayaan/api/import/suppliers", {"content_b64": xlsx, "commit": True})
        self.assertIn("error", response, msg=str(response))
        self.assertFalse(self.env["res.partner"].sudo().search([("ref", "=", "SUP-BLOCK")]))
